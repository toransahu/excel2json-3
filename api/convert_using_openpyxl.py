# -*- coding: utf-8 -*-
"""
~~~~~EXPERIMENTAL~~~~~~
Created on Sat Feb 30 13:49:20 2018
Download & convert a MS Excel sheet to a JSON file [using openpyxl python module, only supports .xlsx files)
"""

import os
import requests
from openpyxl import load_workbook # python module 'openpyxl supports .xlsx only
import json


def download_file(url):
    """
    Download a file from url as raw.

    :param url: url string
    :return: http response
    """
    res = requests.get(url)
    return res


def save_file(loc, filename, response):
    """
    Save a raw file into disk.

    :param loc: directory path string (to save files)
    :param filename: excel file name string
    :param response: http response from url
    :return: void
    """

    file = os.path.join(loc, filename)
    with open(file, 'wb') as f:
        f.write(response.content)


def read_excel_sheet(abspath, sheet_name):
    """
    Read the excel sheet and return it.

    :param abspath: absolute path string to excel file
    :param sheet_name: target excel sheet name string
    :return: worksheet object
    """
    workbook = load_workbook(abspath, data_only=True)
    worksheet = workbook[sheet_name]
    return worksheet


def sheet_to_json(worksheet, loc, sheet_name):
    """
    Read excel sheet and save whole sheet as json file.

    :param worksheet: worksheet object
    :param loc: directory path string (to save files)
    :param sheet_name: sheet_name string
    :return: void
    """
    dict_list = []
    header = [cell[0].value for cell in worksheet.columns]
    for row in worksheet.iter_rows():
        row_dict = {col: row[col_idx].internal_value for col_idx, col in enumerate(header)}
        dict_list.append(row_dict)

    json_file = os.path.join(loc, sheet_name + '.json')
    with open(json_file, 'w') as json_fd:
        json.dump(dict_list, json_fd, sort_keys=False, indent=4)


def main():
    url = 'https://www.iso20022.org/sites/default/files/ISO10383_MIC/ISO10383_MIC.xls'
    loc = '../data/'
    filename = 'ISO10383_MIC.xlsx'
    abspath = os.path.join(loc, filename)
    sheet_name = 'MICs List by CC'

    if not os.path.exists(abspath):
        res = download_file(url)
        save_file(loc, filename, res)
    worksheet = read_excel_sheet(abspath, sheet_name)
    sheet_to_json(worksheet, loc, sheet_name)


main()
